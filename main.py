import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader

class Salary:
    """
    Класс для представления зарплаты

    Attributes:
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вилки оклада
        salary_gross (bool): С учётом налога или нет
        salary_currency (str): Валюта оклада
    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """
        Инициализирует объект Salary

        Args:
            salary_from (str or int or float): Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница вилки оклада
            salary_gross (bool): С учётом налога или нет
            salary_currency (str): Валюта оклада
        """
        self.salary_from = int(float(salary_from))
        self.salary_to = int(float(salary_to))
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def convert_to_rub(self):
        """
        Конвертирует среднюю зарплату в рубли с помощью словаря currency_to_rub

        Returns:
            float: Средняя зарплата в рублях
        """
        return (self.salary_from + self.salary_to) / 2 * self.__currency_to_rub[self.salary_currency]  

    __currency_to_rub = {  
        "AZN": 35.68,  
        "BYR": 23.91,  
        "EUR": 59.90,  
        "GEL": 21.74,  
        "KGS": 0.76,  
        "KZT": 0.13,  
        "RUR": 1,  
        "UAH": 1.64,  
        "USD": 60.66,  
        "UZS": 0.0055,  
    }       

class DataVacancy:
    """
    Класс для представления вакансии

    Attributes:
        name (str): Название вакансии
        salary (Salary): Зарплата вакансии
        area_name (str): Регион вакансии
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        """
        Инициализирует объект DataVacancy

        Args:
            name (str): Название вакансии
            salary_from (str or int or float): Нижняя граница вилки оклада
            salary_to (str or int or float): Верхняя граница вилки оклада
            salary_currency (str): Валюта оклада
            area_name (str): Регион вакансии
            published_at (str): Дата публикации вакансии
        """
        self.name = name
        self.salary = Salary(salary_from, salary_to, False, salary_currency)
        self.area_name = area_name
        self.published_at = published_at

class InputConect:
    """
    Класс для представления входных данных

    Attributes:

    """
    def input_data(self):
        """
        Returns:
            job (str): название выбранной вакансии,
            salary_rub (dict): словарь год - средняя зарплата в рублях,
            salary_count (dict): словарь год - количество вакансий,
            job_rub (dict): словарь год - зарплата выбранной вакансии,
            job_count (dict): словарь год - количество выбранных вакансий ,
            city_salary (dict): словарь город - зарплата,
            city_frac (dict): словарь город - доля выбранных вакансий
        """
        f = input('Введите название файла: ')

        job = input('Введите название профессии: ')

        salary_rub = self.__get_dict()
        salary_count = self.__get_dict()

        job_rub = self.__get_dict()
        job_count = self.__get_dict()

        data_objs = []

        with open(f, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            head = []

            is_first = True
            for row in reader:
                if is_first:  
                    is_first = False
                    head = row
                else:
                    if not "" in row and len(row) == len(head):

                        obj = DataVacancy(
                        row[head.index('name')], 
                        row[head.index('salary_from')], 
                        row[head.index('salary_to')], 
                        row[head.index('salary_currency')], 
                        row[head.index('area_name')], 
                        row[head.index('published_at')]
                        )
                        data_objs.append(obj)

                        year = int(obj.published_at[:4])
                        salary_rub[year] = self.__medium(salary_rub[year], obj.salary.convert_to_rub(), salary_count[year]) 
                        salary_count[year] += 1

                        if(obj.name.find(job) != -1):
                            job_rub[year] = self.__medium(job_rub[year], obj.salary.convert_to_rub(), job_count[year]) 
                            job_count[year] += 1

        salary_rub = self.__erase_empty(self.__round_values(salary_rub))
        job_rub = self.__erase_empty(self.__round_values(job_rub))
        salary_count = self.__erase_empty(salary_count)
        job_count = self.__erase_empty(job_count)

        print('Динамика уровня зарплат по годам:', salary_rub)
        print('Динамика количества вакансий по годам:', salary_count)
        print('Динамика уровня зарплат по годам для выбранной профессии:', job_rub)
        print('Динамика количества вакансий по годам для выбранной профессии:', job_count)
        
        city_salary = {}
        city_count = {}
        city_frac = {} 

        for it in data_objs:
            city = it.area_name
            if city not in city_salary.keys():
                if len([x for x in data_objs if x.area_name == city]) >= int(len(data_objs) / 100):
                    city_salary[city] = it.salary.convert_to_rub()
                    city_count[city] = 1
            else:
                city_salary[city] = self.__medium(city_salary[city], it.salary.convert_to_rub(), city_count[city])
                city_count[city] += 1

        all = len(data_objs)
        for key, value in city_count.items():
            city_frac[key] = round(value / (all / 100) / 100, 4)

        city_salary = self.__round_values(self.__erase_empty(self.__sort_city(city_salary)))
        city_frac = self.__erase_empty(self.__sort_city(city_frac))
        
        print('Уровень зарплат по городам (в порядке убывания):', city_salary)
        print('Доля вакансий по городам (в порядке убывания):', city_frac)

        return job, salary_rub, salary_count, job_rub, job_count, city_salary, city_frac

    def __get_dict(self):
        """
        Возвращает пустой словарь с 2007 года по 2022

        Returns:
            dict: Пустой словарь с 2007 года по 2022
        """
        return {x: 0 for x in range(2007, 2023)}

    def __medium(self, m, x, n):
        """
        Возвращает среднее арифметическое ряда чисел

        Args:
            m (float): Текущая сумма
            x (int): Прибавляемое число
            n (int): Количество уже прибавленных чисел

        Returns:
            float: Среднее арифметическое чисел
        """
        return (m * n + x) / (n + 1)  

    def __sort_city(self, d):
        """
        Сортирует города по зарплате и возвращает первые 10 городов

        Args:
            d (dict): Словарь городов

        Returns:
            dict: Отсортированный по возрастанию словарь городов и зарплат
        """
        return dict(sorted(d.items(), key=lambda x: x[1], reverse=True)[:10])

    def __round_values(self, d):
        """
        Округляет зарплату до целого числа

        Args:
            d (dict): Словарь городов

        Returns:
            dict: Словарь городов с округлённой зарплатой
        """
        return dict(map(lambda x: (x[0], int(x[1])), d.items()))

    def __erase_empty(self, d):
        """
        Удаляет пустые города в словаре

        Args:
            d (dict): словарь городов

        Returns:
            dict: Словарь городов
        """
        cd = dict(filter(lambda x:x[1], d.items()))
        if len(cd.keys()) == 0:
            cd[2022] = 0
        return cd
        
class Report:
    """
    Класс для представления отчёта

    Attributes:
        wb (Workbook): Объект для работы с таблицей эксель
        first_headers (list): Заголовки для первого листа таблицы
        second_headers (list): Заголовки для второго листа таблицы
    """
    def __init__(self):
        """
        Инициализирует объект Report
        """
        self.wb = Workbook()

        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)

        self.wb.create_sheet('Статистика по годам')
        self.wb.create_sheet('Статистика по городам')

    __first_headers = [
        'Год', 
        'Средняя зарплата', 
        'Средняя зарплата - ', 
        'Количество вакансий',
        'Количество вакансий - '
        ]

    def __as_text(self, value):
        """
        Проверяет, является ли объект текстом, если нет - то конвертирует его в строку
        Если невозможно сконвертировать, то возвращает пустую строку

        Args:
            value (object): объект для проверки

        Returns:
            str: Конвертируемый в строку объект
        """
        if value is None:
            return ""
        return str(value)

    def __set_size(self):
        """
        Задаёт размеры колонок в таблице эксель
        """
        for column_cells in self.wb.active.columns:
            length = max(len(self.__as_text(cell.value)) for cell in column_cells)
            self.wb.active.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    def __make_border(self):
        """
        Задаёт обводку для ячеек в таблице эксель
        """
        for row in self.wb.active.rows:
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    def __make_first_sheet(self, data):
        """
        Создаёт первую страницу в эксель и заполняет её

        Args:
            data list(dict): Список словарей со статистикой по годам:
                Динамика уровня зарплат по годам
                Динамика количества вакансий по годам
                Динамика уровня зарплат по годам для выбранной профессии
                Динамика количества вакансий по годам для выбранной профессии

        """
        self.wb.active = self.wb['Статистика по годам']
        ws = self.wb.active
        
        self.__first_headers[2] = self.__first_headers[2] + data[0]
        self.__first_headers[4] = self.__first_headers[4] + data[0]
        ws.append(it for it in self.__first_headers)
        for row in ws.rows:
            for cell in row:
                cell.font = Font(bold=True)
        
        for year in data[1].keys():
            row = [year, data[1][year], data[3][year], data[2][year], data[4][year]]
            ws.append(row)

        self.__set_size()
        self.__make_border()
        
    __second_headers = [
        'Город',
        'Уровень зарплат',
        '',
        'Город',
        'Доля вакансий'
    ]

    def __make_second_sheet(self, data):
        """
        Создаёт вторую странциу в эксель и заполняет её

        Args:
            data list(dict): Список словарей со статистикой по городам
                Уровень зарплат по городам (в порядке убывания)
                Доля вакансий по городам (в порядке убывания)
        """
        self.wb.active = self.wb['Статистика по городам']
        ws = self.wb.active

        ws.append(it for it in self.__second_headers)
        for row in ws.rows:
            for cell in row:
                cell.font = Font(bold=True)
        
        info1 = list(data[0].keys())
        info2 = list(data[0].values())
        info3 = list(data[1].keys())
        info4 = list(data[1].values())

        for i in range(len(data[0])):
            row = [info1[i], info2[i], '', info3[i], info4[i]]
            ws.append(row)

        self.__set_size()
        self.__make_border()

        for i in range(1, 12):
            ws[f"C{i}"].border = Border()
        
        for i in range(1, 12):
            ws[f"E{i}"].number_format = '0.00%'

        self.wb.active = self.wb['Статистика по годам']

    def generate_excel(self, data1, data2):
        """
        Генерирует файл эксель со статистикой

        Args:
            data1 list(dict): Словари для заполнения первой страницы эксель
            data2 list(dict): Словари для заполнения второй страницы эксель
        """
        self.__make_first_sheet(data1)
        self.__make_second_sheet(data2)

        self.wb.save('report.xlsx')

    def __make_salary_year(self, job, data1, data2, ax):
        """
        Создаёт первый график "Уровень зарплат по годам"

        Args:
            job (str): Название вакансии
            data1 (dict): Статистика по всем вакансиям
            data2 (dict): Статистика по выбранной вакансии
            ax (subplot): Объект, куда рисовать график
        """
        labels = list(data1.keys())
        average = list(data1.values())
        jobs = list(data2.values())

        x = np.arange(len(labels))
        width = 0.35
        
        ax.bar(x - width/2, average, width, label='средняя з/п')
        ax.bar(x + width/2, jobs, width, label=f"з/п {job}")    

        ax.set_title('Уровень зарплат по годам')
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(prop={"size":8})
        ax.grid(axis='y')
        ax.tick_params(axis='both', labelsize=8)

    def __make_counts_year(self, job, data1, data2, ax):
        """
        Создаёт второй график "Количество вакансий по годам"

        Args:
            job (str): Название вакансии
            data1 (dict): Статистика по всем вакансиям
            data2 (dict): Статистика по выбранной вакансии
            ax (subplot): Объект, куда рисовать график
        """

        labels = list(data1.keys())
        counts = list(data1.values())
        jobs = list(data2.values())

        x = np.arange(len(labels))
        width = 0.35
        
        ax.bar(x - width/2, counts, width, label='Количество вакансий')
        ax.bar(x + width/2, jobs, width, label=f"Количество вакансий\n{job}")    

        ax.set_title('Количество вакансий по годам')
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(prop={"size":8})
        ax.grid(axis='y')
        ax.tick_params(axis='both', labelsize=8)

    def __make_salary_city(self, data, ax):
        """
        Создаёт третий график "Уровень зарплат по городам"

        Args:
            data (dict): Словарь зарплат по городам
            ax (subplot): Объект, куда рисовать график
        """
        sep = lambda x: x.replace(' ', '\n').replace('-', '\n')

        cities = list(map(sep, data.keys()))[::-1]
        values = list(data.values())[::-1]
        y_pos = np.arange(len(cities))

        ax.barh(y_pos, values)
        ax.set_yticks(y_pos, labels=cities, fontsize=6)
        ax.set_title('Уровень зарплат по городам')
        ax.tick_params(axis='x', labelsize=8)

    def __make_jobs_count(self, data, ax):
        """
        Создаёт четвёртый график "Доля вакансий по городам"

        Args:
            data (dict): Словарь долей вакансий по городам
            ax (subplot): Объект, куда рисовать график
        """
        x = list(data.values())
        x.append(1 - sum(x))
        cities = list(data.keys()) + ['Другие']

        ax.set_title('Доля вакансий по городам')
        ax.pie(x, labels = cities, textprops={'fontsize': 6}, startangle=90)

    def generate_image(self, data):
        """
        Создаёт графическую статистику

        Args:
            data list(dict) - Список словарей со статистикой
                Динамика уровня зарплат по годам
                Динамика количества вакансий по годам
                Динамика уровня зарплат по годам для выбранной профессии
                Динамика количества вакансий по годам для выбранной профессии
                Уровень зарплат по городам (в порядке убывания)
                Доля вакансий по городам (в порядке убывания)
        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        self.__make_salary_year(data[0], data[1], data[3], ax1)
        self.__make_counts_year(data[0], data[2], data[4], ax2)
        self.__make_salary_city(data[5], ax3)
        self.__make_jobs_count(data[6], ax4)
        
        fig.tight_layout()
        fig.savefig('graph.png')
        
    def generate_pdf(self, data):
        """
        Создаёт pdf статистику

        Args:
            data list(dict) - Список словарей со статистикой
                Динамика уровня зарплат по годам
                Динамика количества вакансий по годам
                Динамика уровня зарплат по годам для выбранной профессии
                Динамика количества вакансий по годам для выбранной профессии
                Уровень зарплат по городам (в порядке убывания)
                Доля вакансий по городам (в порядке убывания)
        """
        job = data[0]
        image_file = "graph.png"

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render({'job': job, 'image_file': image_file})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config)

ic = InputConect()
data = ic.input_data()

data = list(data)
